import os
import openpyxl


# print out currrent directory
mypwd = os.getcwd()
print("current dir 1= ", mypwd)

# change directory
#os.chdir("C:/Users/mrkdf/Desktop/testPython")
mypwd = os.getcwd()
print("current dir 2= ", mypwd)

# myfile = "C:/Users/mrkdf/Desktop/testPython/example1.xlsx"
# wb = openpyxl.load_workbook(myfile)

# myfile2 = 'Users/mrkdf/Desktop/testPython/example1.xlsx'
# wb = openpyxl.load_workbook(myfile2)

myfile3 = "example1.xlsx"
wb = openpyxl.load_workbook(myfile3)

sheet = wb.get_sheet_by_name('Sheet1')
type(sheet)
a1=sheet['A1'].value
print('a1= ', a1)
b1=sheet['B1'].value
print('b1= ', b1 )
ma_row = sheet.max_row
ma_col = sheet.max_column
print("ma_row =  ", ma_row)
print("ma_col = ", ma_col)
print("------- done -------")